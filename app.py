"""
Seating Arrangement Generator - 1st Summative Evaluation 2026
Streamlit app to auto-generate exam seating from Excel student data.
"""

import streamlit as st
import pandas as pd
import io
import math
from collections import defaultdict

import openpyxl
import openpyxl.worksheet.pagebreak
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak, HRFlowable
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.graphics import renderPDF

# ─── Constants & Helpers ─────────────────────────────────────────────────────

CLASS_ORDER = ["V", "VI", "VII", "VIII", "IX", "X"]

CLASS_COLORS = {
    "V":    "#4CAF50",
    "VI":   "#2196F3",
    "VII":  "#9C27B0",
    "VIII": "#FF5722",
    "IX":   "#00BCD4",
    "X":    "#FF9800",
}

SCHOOL_NAME = "1st Summative Evaluation 2026"


def sort_classes(class_iterable):
    valid_classes = [c for c in class_iterable if pd.notna(c)]
    return sorted(valid_classes, key=lambda c: CLASS_ORDER.index(c) if c in CLASS_ORDER else 999)


# ─── Data Ingestion ──────────────────────────────────────────────────────────

def _normalize_class(raw):
    s = str(raw).strip()
    for prefix, cls in [
        ("V  : ", "V"), ("VI : ", "VI"), ("VII : ", "VII"),
        ("VIII : ", "VIII"), ("IX", "IX"), ("X", "X"),
    ]:
        if s.startswith(prefix):
            return cls
    return None


def read_students(file) -> pd.DataFrame:
    df = pd.read_excel(file, sheet_name="Sheet1", header=None)
    rows = []
    for _, r in df.iloc[2:].iterrows():
        cls = _normalize_class(r[0])
        if cls is None:
            continue
        try:
            roll = int(r[1])
        except (ValueError, TypeError):
            continue
        gender = str(r[10]).strip().upper()
        if gender not in ("BOYS", "GIRLS"):
            continue
        name = str(r[4]).strip() if pd.notna(r[4]) else ""
        rows.append({"class": cls, "roll": roll, "name": name, "gender": gender})

    df_out = pd.DataFrame(rows)
    if not df_out.empty:
        df_out['class_rank']  = df_out['class'].apply(lambda x: CLASS_ORDER.index(x) if x in CLASS_ORDER else 99)
        df_out['gender_rank'] = df_out['gender'].apply(lambda x: 0 if x == 'BOYS' else 1)
        df_out = df_out.sort_values(['class_rank', 'gender_rank', 'roll']).drop(columns=['class_rank', 'gender_rank']).reset_index(drop=True)
    return df_out


# ─── Dynamic Room Distribution ───────────────────────────────────────────────

def distribute_to_rooms(df: pd.DataFrame, rooms_config: list, separate_genders: bool) -> tuple[dict, list]:
    allocated_rooms    = {r["name"]: [] for r in rooms_config}
    unassigned_students = []
    room_capacities    = {r["name"]: sum(r["cols"]) * 3 for r in rooms_config}
    room_gender_locks  = {r["name"]: None for r in rooms_config}

    def pop_mixed_student(student_dict):
        available_classes = [c for c in student_dict if student_dict[c]]
        if not available_classes:
            return None
        available_classes.sort(key=lambda c: len(student_dict[c]), reverse=True)
        return student_dict[available_classes[0]].pop(0)

    queues = {"BOYS": defaultdict(list), "GIRLS": defaultdict(list)}
    for s in df.to_dict("records"):
        queues[s["gender"]][s["class"]].append(s)

    genders_to_process = ["BOYS", "GIRLS"] if separate_genders else ["MIXED"]

    if not separate_genders:
        mixed_queues = defaultdict(list)
        for g in ["BOYS", "GIRLS"]:
            for c, students in queues[g].items():
                mixed_queues[c].extend(students)
        queues = {"MIXED": mixed_queues}

    for target_gender in genders_to_process:
        active_queue = queues[target_gender]
        while any(active_queue.values()):
            student = pop_mixed_student(active_queue)
            if not student:
                break
            placed = False
            for room in rooms_config:
                r_name = room["name"]
                if len(allocated_rooms[r_name]) >= room_capacities[r_name]:
                    continue
                if separate_genders:
                    current_lock = room_gender_locks[r_name]
                    if current_lock is None:
                        room_gender_locks[r_name] = target_gender
                    elif current_lock != target_gender:
                        continue
                allocated_rooms[r_name].append(student)
                placed = True
                break
            if not placed:
                unassigned_students.append(student)

    return allocated_rooms, unassigned_students


# ─── Bench Seating Algorithm ─────────────────────────────────────────────────

def create_bench_layout(students: list[dict]) -> list[list]:
    groups = defaultdict(list)
    for s in students:
        groups[s["class"]].append(s)

    class_order = sort_classes(groups.keys())
    queues = {c: list(groups[c]) for c in class_order}
    benches = []

    while any(queues.values()):
        available = [(c, queues[c]) for c in class_order if queues[c]]
        if not available:
            break
        if len(available) == 1:
            cls, q = available[0]
            while q:
                triple = [q.pop(0), q.pop(0) if q else None, q.pop(0) if q else None]
                benches.append(triple)
            break
        available.sort(key=lambda x: len(x[1]), reverse=True)
        cls_a, q_a = available[0]
        cls_b, q_b = available[1]
        left   = q_a.pop(0)
        middle = q_b.pop(0)
        right  = q_a.pop(0) if q_a else None
        benches.append([left, middle, right])

    return benches


# ─── PDF Generation ──────────────────────────────────────────────────────────

def _style(name, **kwargs):
    base = dict(fontName="Helvetica", fontSize=9, alignment=TA_CENTER)
    base.update(kwargs)
    return ParagraphStyle(name, **base)


def _seat_cell(student):
    if student is None:
        return "—"
    g = "Boy" if student["gender"] == "BOYS" else "Girl"
    return f"Roll: {student['roll']}\nClass {student['class']}  [{g}]\n{student['name']}"


def _room_diagram(benches: list[list], room_config: dict) -> Drawing:
    col_heights = room_config["cols"]
    B_W, B_H   = 50, 34
    GAP_X, GAP_Y = 8, 8
    SEAT_R     = 6
    COLS       = len(col_heights)
    max_rows   = max(col_heights) if col_heights else 1
    dw         = COLS * (B_W + GAP_X) + GAP_X
    dh         = max_rows * (B_H + GAP_Y) + GAP_Y + 22
    d          = Drawing(dw, dh)

    board_w = min(dw * 0.6, 200)
    bx = (dw - board_w) / 2
    d.add(Rect(bx, dh - 20, board_w, 14, fillColor=colors.HexColor("#2e7d32"), strokeColor=colors.HexColor("#1b5e20"), strokeWidth=1))
    d.add(String(dw / 2, dh - 13, "BLACKBOARD", fontName="Helvetica-Bold", fontSize=7, fillColor=colors.white, textAnchor="middle"))

    bench_idx = 0
    for col_idx, rows_in_col in enumerate(col_heights):
        for row_idx in range(rows_in_col):
            if bench_idx >= len(benches):
                break
            bench = benches[bench_idx]
            x = GAP_X + col_idx * (B_W + GAP_X)
            y = dh - 22 - (row_idx + 1) * (B_H + GAP_Y)
            cls      = next((s["class"] for s in bench if s), "V")
            fill_hex = CLASS_COLORS.get(cls, "#90caf9")
            fill     = colors.HexColor(fill_hex)
            d.add(Rect(x, y, B_W, B_H, fillColor=colors.HexColor("#e3f2fd"), strokeColor=colors.HexColor("#90caf9"), strokeWidth=0.8))
            d.add(String(x + B_W / 2, y + B_H - 9, f"B{bench_idx+1}", fontName="Helvetica-Bold", fontSize=6, fillColor=colors.HexColor("#1a237e"), textAnchor="middle"))
            for sp_x, sp_y in [(x + 10, y + 10), (x + B_W / 2, y + 10), (x + B_W - 10, y + 10)]:
                d.add(Rect(sp_x - SEAT_R, sp_y - SEAT_R, SEAT_R * 2, SEAT_R * 2, fillColor=fill, strokeColor=colors.HexColor("#37474f"), strokeWidth=0.6))
            bench_idx += 1
    return d


def generate_pdf(allocated_rooms: dict, rooms_config: list) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=13*mm, rightMargin=13*mm, topMargin=13*mm, bottomMargin=13*mm)
    st_exam    = _style("exam",    fontName="Helvetica-Bold", fontSize=15, spaceAfter=1*mm,  textColor=colors.HexColor("#0d1b2a"))
    st_room    = _style("room",    fontName="Helvetica-Bold", fontSize=22, spaceAfter=3*mm,  textColor=colors.HexColor("#0f3460"))
    st_section = _style("section", fontName="Helvetica-Bold", fontSize=9,  spaceAfter=2*mm,  textColor=colors.HexColor("#37474f"))
    st_footer  = _style("footer",  fontSize=7, textColor=colors.grey, fontName="Helvetica-Oblique")
    story = []

    for idx, config in enumerate(rooms_config):
        r_name   = config["name"]
        students = allocated_rooms[r_name]
        if not students:
            continue
        if idx > 0:
            story.append(PageBreak())

        benches = create_bench_layout(students)
        story.append(Paragraph(SCHOOL_NAME, st_exam))
        story.append(Paragraph(str(r_name).upper(), st_room))
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#0f3460"), spaceAfter=3*mm))

        counts = defaultdict(lambda: {"BOYS": 0, "GIRLS": 0})
        for s in students:
            counts[s["class"]][s["gender"]] += 1

        hdr = [["Class", "Boys", "Girls", "Total"]]
        body, tb, tg = [], 0, 0
        for cls in sort_classes(counts.keys()):
            b, g = counts[cls]["BOYS"], counts[cls]["GIRLS"]
            body.append([f"Class {cls}", str(b) if b else "–", str(g) if g else "–", str(b + g)])
            tb += b; tg += g
        body.append(["TOTAL", str(tb), str(tg), str(tb + tg)])

        summary_tbl = Table(hdr + body, colWidths=[35*mm, 24*mm, 24*mm, 24*mm], hAlign="CENTER")
        summary_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f3460")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#dde8f0")),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#aaaaaa")),
        ]))

        story.append(Paragraph("CLASS-WISE STUDENT COUNT", st_section))
        story.append(summary_tbl)
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(f"CLASSROOM LAYOUT  (Total Benches: {sum(config['cols'])})", st_section))
        story.append(_room_diagram(benches, config))
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph("BENCH-WISE SEATING ARRANGEMENT", st_section))

        bench_hdr  = [["Bench\nNo.", "LEFT SEAT\n(Roll | Class | Gender | Name)", "MIDDLE SEAT\n(Roll | Class | Gender | Name)", "RIGHT SEAT\n(Roll | Class | Gender | Name)"]]
        bench_rows = [[str(i + 1)] + [_seat_cell(s) for s in bench] for i, bench in enumerate(benches)]
        bench_tbl  = Table(bench_hdr + bench_rows, colWidths=[12*mm, 53*mm, 53*mm, 53*mm], repeatRows=1)
        bench_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16213e")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#e8eaf6")),
            ("FONTSIZE",   (0, 0), (-1, -1), 7.5),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#c0c0c0")),
        ]))
        story.append(bench_tbl)
        story.append(Spacer(1, 3*mm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#aaaaaa")))
        story.append(Paragraph(f"{r_name}  ·  Total Students: {len(students)}  ·  {SCHOOL_NAME}", st_footer))

    doc.build(story)
    buf.seek(0)
    return buf


# ─── Excel Generation (Exact Format Match + Auto Page Header Repeat) ─────────

# Max data rows per printed page before headers must repeat
ROWS_PER_PAGE = 29


def generate_student_list_excel(student_df, classes_to_print):
    """
    Generates an Excel workbook with one sheet per class.

    Layout exactly matches Student_List_Selected_Class_Only.xlsx:
      Col A  = SL. (Boys)         width  9.14
      Col B  = Roll No (Boys)     width  default (~8.43)
      Col C  = Name (Boys)        width 26.00
      Col D  = SPACER (blank)     width 29.29  ← no borders, no data
      Col E  = SL. (Girls)        width  9.14
      Col F  = Roll No (Girls)    width 10.29
      Col G  = Name (Girls)       width 30.29

    Pagination logic:
      - Each printed page fits exactly ROWS_PER_PAGE (29) data rows.
      - If boys OR girls count exceeds 29, a fresh pair of gender headers +
        sub-headers is automatically inserted at the start of every new page
        block so the printed output always has headers at the top of each page.
      - Boys and girls are chunked independently and paired chunk-by-chunk.

    Page   : A4 Landscape
    Margins: Left 0.75" · Right 0.75" · Top 1.0" · Bottom 1.0" · Header 0.5" · Footer 0.5"
    Font   : Calibri 8 pt  (all cells, not bold)
    Border : Thin all sides on A–C and E–G; Column D has no border
    Align  : Center-Center on all cells
    """

    # ── Shared style objects ─────────────────────────────────────────────────
    FONT        = Font(name="Calibri", size=8)
    ALIGN_CC    = Alignment(horizontal="center", vertical="center")
    ALIGN_WRAP  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN        = Side(style="thin")
    BORDER_ALL  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BORDER_NONE = Border()

    def apply(cell, value, alignment=None, border=None):
        cell.value     = value
        cell.font      = FONT
        cell.alignment = alignment or ALIGN_CC
        cell.border    = border    or BORDER_ALL

    def write_headers(ws, header_row, cls):
        """Write the two-row header block (gender titles + sub-headers) at header_row."""
        sub_row = header_row + 1

        # ── Gender title row ─────────────────────────────────────────────
        ws.merge_cells(f"A{header_row}:C{header_row}")
        apply(ws[f"A{header_row}"], f"GENDER - MALE (Class {cls})")

        ws.merge_cells(f"E{header_row}:G{header_row}")
        apply(ws[f"E{header_row}"], f"GENDER - FEMALE (Class {cls})")

        ws[f"D{header_row}"].border = BORDER_NONE

        # ── Sub-header row (height 15 pt) ────────────────────────────────
        ws.row_dimensions[sub_row].height = 15.0

        for col, label in [("A", "SL."), ("B", "Roll Number"), ("C", "Student Name")]:
            apply(ws[f"{col}{sub_row}"], label,
                  alignment=ALIGN_WRAP if label == "Roll Number" else ALIGN_CC)

        for col, label in [("E", "SL."), ("F", "Roll Number"), ("G", "Student Name")]:
            apply(ws[f"{col}{sub_row}"], label,
                  alignment=ALIGN_WRAP if label == "Roll Number" else ALIGN_CC)

        ws[f"D{sub_row}"].border = BORDER_NONE

    def write_data_block(ws, boys_chunk, girls_chunk, data_start_row,
                         boys_sl_offset, girls_sl_offset):
        """
        Write one chunk of boys + girls data starting at data_start_row.
        sl_offset = number of rows already written in previous chunks
                    (so SL. numbers continue from where they left off).
        """
        chunk_len = max(len(boys_chunk), len(girls_chunk))
        for i in range(chunk_len):
            row = data_start_row + i

            # Boys side ──────────────────────────────────────────────────
            if i < len(boys_chunk):
                apply(ws[f"A{row}"], boys_sl_offset + i + 1)
                apply(ws[f"B{row}"], boys_chunk.iloc[i]["roll"])
                apply(ws[f"C{row}"], boys_chunk.iloc[i]["name"])
            else:
                for col in ("A", "B", "C"):
                    ws[f"{col}{row}"].border = BORDER_ALL
                    ws[f"{col}{row}"].font   = FONT

            # Spacer ─────────────────────────────────────────────────────
            ws[f"D{row}"].border = BORDER_NONE

            # Girls side ─────────────────────────────────────────────────
            if i < len(girls_chunk):
                apply(ws[f"E{row}"], girls_sl_offset + i + 1)
                apply(ws[f"F{row}"], girls_chunk.iloc[i]["roll"])
                apply(ws[f"G{row}"], girls_chunk.iloc[i]["name"])
            else:
                for col in ("E", "F", "G"):
                    ws[f"{col}{row}"].border = BORDER_ALL
                    ws[f"{col}{row}"].font   = FONT

        return chunk_len   # number of data rows written

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for cls in classes_to_print:
        cls_data = student_df[student_df["class"] == cls].copy()
        if cls_data.empty:
            continue

        boys  = cls_data[cls_data["gender"] == "BOYS" ].sort_values("roll").reset_index(drop=True)
        girls = cls_data[cls_data["gender"] == "GIRLS"].sort_values("roll").reset_index(drop=True)

        ws = wb.create_sheet(title=f"Class {cls}")

        # ── Page setup ───────────────────────────────────────────────────────
        ws.page_setup.paperSize   = 9        # A4
        ws.page_setup.orientation = "landscape"
        ws.page_margins = PageMargins(
            left=0.75,  right=0.75,
            top=1.0,    bottom=1.0,
            header=0.5, footer=0.5
        )

        # ── Default row height ────────────────────────────────────────────────
        ws.sheet_format.defaultRowHeight = 11.25
        ws.sheet_format.customHeight     = True

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 9.140625
        # Col B left at Excel default (~8.43)
        ws.column_dimensions["C"].width = 26.0
        ws.column_dimensions["D"].width = 29.28515625
        ws.column_dimensions["E"].width = 9.140625
        ws.column_dimensions["F"].width = 10.28515625
        ws.column_dimensions["G"].width = 30.28515625

        # ── Chunk boys and girls into ROWS_PER_PAGE slices ───────────────────
        def chunks(df, size):
            return [df.iloc[i:i + size].reset_index(drop=True)
                    for i in range(0, max(len(df), 1), size)]

        boys_chunks  = chunks(boys,  ROWS_PER_PAGE)
        girls_chunks = chunks(girls, ROWS_PER_PAGE)

        # Pad shorter list with empty DataFrames so both have same length
        num_pages = max(len(boys_chunks), len(girls_chunks))
        empty_df  = pd.DataFrame(columns=boys.columns)
        while len(boys_chunks)  < num_pages: boys_chunks.append(empty_df)
        while len(girls_chunks) < num_pages: girls_chunks.append(empty_df)

        # ── Write each page block ─────────────────────────────────────────────
        current_row   = 1   # next Excel row to write into
        boys_written  = 0
        girls_written = 0

        for page_idx in range(num_pages):
            b_chunk = boys_chunks[page_idx]
            g_chunk = girls_chunks[page_idx]

            # Write 2-row header block
            write_headers(ws, current_row, cls)
            data_start = current_row + 2          # data begins after 2 header rows

            # Set a manual page break before every page block except the first
            if page_idx > 0:
                ws.row_breaks.append(
                    openpyxl.worksheet.pagebreak.Break(id=current_row - 1)
                )

            # Write data rows for this chunk
            rows_written = write_data_block(
                ws, b_chunk, g_chunk,
                data_start,
                boys_written,
                girls_written
            )

            boys_written  += len(b_chunk)
            girls_written += len(g_chunk)
            current_row    = data_start + rows_written   # advance cursor

    # ── Save & return ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Streamlit UI ─────────────────────────────────────────────────────────────

def main():
    st.set_page_config(page_title="Seating Arrangement Generator", page_icon="🏫", layout="wide")

    st.markdown("""
    <style>
        .main-title { font-size:2.2rem; font-weight:800; color:#0f3460; margin-bottom:0; }
        .sub-title  { font-size:1rem;   color:#555;      margin-bottom:1.5rem; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<p class="main-title">🏫 Seating Arrangement Generator</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-title">Fully Configurable Automated PDF & Excel Generation</p>', unsafe_allow_html=True)

    with st.sidebar:
        st.header("⚙️ Configuration")
        class_mode = st.radio("Class Selection", ["All Classes", "Custom Classes"])
        st.markdown("---")
        st.subheader("Gender Rules")
        separate_genders = st.checkbox("🚫 Separate Boys & Girls into different rooms", value=False)

    uploaded = st.file_uploader("📂 Upload Student Excel File (.xlsx)", type=["xlsx"])

    if not uploaded:
        st.info("👆 Please upload the Excel file to begin.")
        return

    with st.spinner("Reading Excel..."):
        try:
            raw_df = read_students(uploaded)
        except Exception as e:
            st.error(f"Error reading file: {e}")
            return

    if raw_df.empty:
        st.warning("No valid students found in the file. Check formatting.")
        return

    available_classes = sort_classes(raw_df["class"].unique())
    if class_mode == "Custom Classes":
        selected_classes = st.sidebar.multiselect("Select Classes to Process", available_classes, default=available_classes)
        df = raw_df[raw_df["class"].isin(selected_classes)]
    else:
        df = raw_df

    st.success(f"✅ Loaded **{len(df):,}** students to process.")

    # =========================================================================
    # CLASS SUMMARY & STUDENT LISTS
    # =========================================================================
    st.markdown("---")
    st.header("📋 Class Summary & Student Lists")

    if not df.empty:
        st.subheader("Class Summary")
        summary_data = []
        for cls in sort_classes(df["class"].unique()):
            cls_df = df[df["class"] == cls]
            boys_count  = len(cls_df[cls_df["gender"] == "BOYS"])
            girls_count = len(cls_df[cls_df["gender"] == "GIRLS"])
            summary_data.append({
                "Class": cls,
                "Total Students": boys_count + girls_count,
                "Boys": boys_count,
                "Girls": girls_count,
            })

        summary_df = pd.DataFrame(summary_data)
        st.dataframe(summary_df, use_container_width=True)

        # ── Summary PDF ──────────────────────────────────────────────────────
        def generate_summary_pdf(sum_df):
            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
            elements   = []
            title_style = ParagraphStyle(name="Title", fontSize=16, alignment=TA_CENTER, fontName="Helvetica-Bold", spaceAfter=10*mm)
            elements.append(Paragraph(f"{SCHOOL_NAME} - Class Summary", title_style))
            data = [["Class", "Total Students", "Boys", "Girls"]]
            for _, row in sum_df.iterrows():
                data.append([str(row["Class"]), str(row["Total Students"]), str(row["Boys"]), str(row["Girls"])])
            t = Table(data, colWidths=[40*mm, 40*mm, 40*mm, 40*mm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f3460")),
                ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
                ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f5f8ff")),
                ("GRID",       (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(t)
            doc.build(elements)
            buf.seek(0)
            return buf

        st.download_button(
            label="📥 Download Class Summary PDF",
            data=generate_summary_pdf(summary_df),
            file_name="Class_Summary.pdf",
            mime="application/pdf",
        )

        st.markdown("<br>", unsafe_allow_html=True)
        st.subheader("Generate Class-wise Student List")

        col1, col2 = st.columns(2)
        with col1:
            list_option = st.radio("Select Generation Mode:", ["All Classes", "Selected Class Only"])

        selected_cls_list = []
        with col2:
            if list_option == "Selected Class Only":
                selected_cls       = st.selectbox("Choose Class", sort_classes(df["class"].unique()))
                selected_cls_list  = [selected_cls]
            else:
                selected_cls_list  = sort_classes(df["class"].unique())

        # ── Student List PDF ─────────────────────────────────────────────────
        def generate_student_list_pdf(student_df, classes_to_print):
            """
            A4 Landscape PDF — one table per class with repeatRows=2.
            ReportLab auto-splits the table across pages and repeats the
            2 header rows (gender title + sub-headers) at the top of every
            new page automatically — no manual chunking needed here.

            Layout:  GENDER - MALE (Class X)  |  gap  |  GENDER - FEMALE (Class X)
                     SL. | Roll Number | Name  |       |  SL. | Roll Number | Name
                     data rows …
            """
            from reportlab.lib.pagesizes import landscape

            A4_L = landscape(A4)          # 841.89pt × 595.28pt  (297mm × 210mm)

            # Margins exactly matching the Excel page setup
            L_MAR = 0.75 * 25.4 * mm     # 0.75 inch
            R_MAR = 0.75 * 25.4 * mm
            T_MAR = 1.0  * 25.4 * mm     # 1.0 inch
            B_MAR = 1.0  * 25.4 * mm

            usable_w = A4_L[0] - L_MAR - R_MAR   # ≈ 733 pt

            # ── Column widths — proportional to Excel character units ─────────
            # Boys: A=9.14  B=8.43  C=26.0  |  gap=8mm  |  Girls: E=9.14  F=10.29  G=30.29
            TOTAL_UNITS = 9.14 + 8.43 + 26.0 + 9.14 + 10.29 + 30.29  # 93.29
            GAP_W       = 8 * mm
            scale       = (usable_w - GAP_W) / TOTAL_UNITS
            CW = [
                9.14  * scale,   # col 0  SL.   (Boys)
                8.43  * scale,   # col 1  Roll  (Boys)
                26.0  * scale,   # col 2  Name  (Boys)
                GAP_W,           # col 3  spacer — no border
                9.14  * scale,   # col 4  SL.   (Girls)
                10.29 * scale,   # col 5  Roll  (Girls)
                30.29 * scale,   # col 6  Name  (Girls)
            ]

            HDR_BG   = colors.HexColor("#dde8f0")
            HDR_FONT = "Helvetica-Bold"
            DAT_FONT = "Helvetica"
            FONT_SZ  = 8
            BDR_CLR  = colors.black
            WHT      = colors.white

            # ── Build PDF ────────────────────────────────────────────────────
            buf  = io.BytesIO()
            doc  = SimpleDocTemplate(
                buf, pagesize=A4_L,
                leftMargin=L_MAR, rightMargin=R_MAR,
                topMargin=T_MAR,  bottomMargin=B_MAR,
            )
            elements     = []
            first_class  = True

            for cls in classes_to_print:
                cls_data = student_df[student_df["class"] == cls].copy()
                if cls_data.empty:
                    continue

                boys  = cls_data[cls_data["gender"] == "BOYS" ].sort_values("roll").reset_index(drop=True)
                girls = cls_data[cls_data["gender"] == "GIRLS"].sort_values("roll").reset_index(drop=True)

                # ── Build row data ────────────────────────────────────────────
                # Row 0: gender title headers (will be repeated by repeatRows=2)
                # Row 1: column sub-headers  (will be repeated by repeatRows=2)
                # Row 2+: data
                rows = [
                    [f"GENDER - MALE (Class {cls})", "", "", "",
                     f"GENDER - FEMALE (Class {cls})", "", ""],
                    ["SL.", "Roll Number", "Student Name", "",
                     "SL.", "Roll Number", "Student Name"],
                ]

                max_len = max(len(boys), len(girls))
                for i in range(max_len):
                    rows.append([
                        str(i + 1)                  if i < len(boys)  else "",
                        str(boys.iloc[i]["roll"])   if i < len(boys)  else "",
                        str(boys.iloc[i]["name"])   if i < len(boys)  else "",
                        "",   # spacer
                        str(i + 1)                  if i < len(girls) else "",
                        str(girls.iloc[i]["roll"])  if i < len(girls) else "",
                        str(girls.iloc[i]["name"])  if i < len(girls) else "",
                    ])

                n = len(rows)   # total rows including both header rows

                # ── Table with repeatRows=2 ───────────────────────────────────
                # repeatRows=2 tells ReportLab: whenever this table is split
                # across pages, repeat the first 2 rows at the top of each
                # continuation page — so headers always appear, no gaps.
                tbl = Table(rows, colWidths=CW, repeatRows=2)

                tbl.setStyle(TableStyle([
                    # Global
                    ("FONTNAME",       (0, 0), (-1, -1), DAT_FONT),
                    ("FONTSIZE",       (0, 0), (-1, -1), FONT_SZ),
                    ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING",     (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING",  (0, 0), (-1, -1), 2),

                    # Row 0 — gender title: span + bold + bg
                    ("SPAN",           (0, 0), (2, 0)),
                    ("SPAN",           (4, 0), (6, 0)),
                    ("FONTNAME",       (0, 0), (2, 0), HDR_FONT),
                    ("FONTNAME",       (4, 0), (6, 0), HDR_FONT),
                    ("BACKGROUND",     (0, 0), (2, 0), HDR_BG),
                    ("BACKGROUND",     (4, 0), (6, 0), HDR_BG),

                    # Row 1 — sub-headers: bold + bg
                    ("FONTNAME",       (0, 1), (2, 1), HDR_FONT),
                    ("FONTNAME",       (4, 1), (6, 1), HDR_FONT),
                    ("BACKGROUND",     (0, 1), (2, 1), HDR_BG),
                    ("BACKGROUND",     (4, 1), (6, 1), HDR_BG),

                    # Borders — Boys side (cols 0-2)
                    ("BOX",            (0, 0), (2, n - 1), 0.5, BDR_CLR),
                    ("INNERGRID",      (0, 0), (2, n - 1), 0.5, BDR_CLR),

                    # Borders — Girls side (cols 4-6)
                    ("BOX",            (4, 0), (6, n - 1), 0.5, BDR_CLR),
                    ("INNERGRID",      (4, 0), (6, n - 1), 0.5, BDR_CLR),

                    # Spacer col 3 — white, zero-width borders
                    ("BACKGROUND",     (3, 0), (3, n - 1), WHT),
                    ("LINEAFTER",      (2, 0), (2, n - 1), 0,   WHT),
                    ("LINEBEFORE",     (4, 0), (4, n - 1), 0,   WHT),

                    # Row heights
                    ("ROWHEIGHT",      (0, 0), (-1, 0),     15),  # gender title
                    ("ROWHEIGHT",      (0, 1), (-1, 1),     15),  # sub-header
                    ("ROWHEIGHT",      (0, 2), (-1, n - 1), 13),  # data rows
                ]))

                if not first_class:
                    elements.append(PageBreak())
                first_class = False

                elements.append(tbl)

            doc.build(elements)
            buf.seek(0)
            return buf

        # ── Download buttons ─────────────────────────────────────────────────
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            st.download_button(
                label="📄 Download Student List (PDF)",
                data=generate_student_list_pdf(df, selected_cls_list),
                file_name=f"Student_List_{list_option.replace(' ', '_')}.pdf",
                mime="application/pdf",
                type="primary",
                use_container_width=True,
            )
        with col_dl2:
            st.download_button(
                label="📊 Download Student List (EXCEL)",
                data=generate_student_list_excel(df, selected_cls_list),
                file_name=f"Student_List_{list_option.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.markdown("---")

    # =========================================================================
    # ROOM CONFIGURATION & SEATING
    # =========================================================================
    st.subheader("🚪 Room Configuration")
    st.markdown("Add rooms and define their seating layout. Enter benches per column separated by commas (e.g., `6,9` = two columns of 6 and 9 benches).")

    default_rooms = pd.DataFrame([
        {"Room Name": "Room 1", "Layout (comma separated)": "11,11"},
        {"Room Name": "Room 2", "Layout (comma separated)": "6,6"},
        {"Room Name": "Room 3", "Layout (comma separated)": "7,6"},
        {"Room Name": "Room 4", "Layout (comma separated)": "6,6"},
        {"Room Name": "Room 5", "Layout (comma separated)": "9,9"},
    ])

    edited_room_df = st.data_editor(default_rooms, num_rows="dynamic", use_container_width=True)

    rooms_config = []
    total_system_capacity = 0
    for _, row in edited_room_df.iterrows():
        name       = str(row["Room Name"]).strip()
        layout_str = str(row["Layout (comma separated)"]).strip()
        if not name or not layout_str:
            continue
        try:
            layout_str = layout_str.replace(":", ",")
            cols = [int(c.strip()) for c in layout_str.split(",") if c.strip().isdigit()]
            if cols:
                room_cap               = sum(cols) * 3
                total_system_capacity += room_cap
                rooms_config.append({"name": name, "cols": cols, "capacity": room_cap})
        except ValueError:
            st.error(f"Invalid layout format in {name}. Please use numbers separated by commas.")
            return

    st.info(f"🪑 **Total System Capacity:** {total_system_capacity} Seats | **Total Students:** {len(df)}")
    if len(df) > total_system_capacity:
        st.error(f"⚠️ Warning: Not enough seats! You are short by {len(df) - total_system_capacity} seats.")

    allocated_rooms, unassigned = distribute_to_rooms(df, rooms_config, separate_genders)

    if unassigned:
        st.error(f"⚠️ {len(unassigned)} students could not be seated due to lack of space or strict gender isolation rules.")
        with st.expander("View Unassigned Students"):
            st.dataframe(unassigned)

    st.subheader("📊 Allocation Preview")
    preview_data = []
    for config in rooms_config:
        r_name   = config["name"]
        students = allocated_rooms[r_name]
        preview_data.append({
            "Room Name":      r_name,
            "Assigned Boys":  sum(1 for s in students if s["gender"] == "BOYS"),
            "Assigned Girls": sum(1 for s in students if s["gender"] == "GIRLS"),
            "Total Occupied": f"{len(students)} / {config['capacity']}",
        })
    st.dataframe(pd.DataFrame(preview_data).set_index("Room Name"), use_container_width=True)

    if st.button("🖨️ Generate Seating Arrangement PDF", type="primary", use_container_width=True):
        with st.spinner("Calculating matrices and rendering PDF..."):
            try:
                pdf_buf = generate_pdf(allocated_rooms, rooms_config)
                st.balloons()
                st.download_button(
                    label="📥 Download Final Seating PDF",
                    data=pdf_buf,
                    file_name="Custom_Seating_Arrangement.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"PDF generation failed: {e}")
                st.exception(e)


if __name__ == "__main__":
    main()
